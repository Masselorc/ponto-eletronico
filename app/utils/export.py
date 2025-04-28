# -*- coding: utf-8 -*-
import os
from datetime import datetime, date
from flask import render_template, current_app
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from app.models.ponto import Ponto, Atividade # Mantém importações de modelos
from app.models.feriado import Feriado
from app.models.user import User
from app.models.relatorio_completo import RelatorioMensalCompleto # Adicionado para generate_pdf
# Removido: from calendar import monthrange (agora em helpers.py)
# --- Importa a função auxiliar do novo módulo ---
from app.utils.helpers import _get_relatorio_mensal_data
# -------------------------------------------------
import logging # Adicionado para logging

logger = logging.getLogger(__name__) # Configura logger

# Função create_pdf (mantida como estava)
def create_pdf(template_name, output_path, **context):
    """Cria um arquivo PDF a partir de um template HTML."""
    try:
        html_content = render_template(template_name, **context)
        # Estilos CSS (mantidos como na sua versão anterior)
        pdf_styles = """
        <style>
            @page {
                size: A4 portrait; /* Define tamanho A4 e orientação retrato */
                margin: 1.5cm 1.5cm 2cm 1.5cm; /* Margens (top, right, bottom, left) - Aumentada inferior para rodapé */

                /* Opcional: Adicionar rodapé diretamente com CSS Paged Media */
                @bottom-center {
                    content: "Página " counter(page) " de " counter(pages);
                    font-size: 8pt;
                    color: #6c757d;
                }
                 @bottom-left {
                    content: "Sistema de Ponto Eletrônico - SENAPPEN";
                    font-size: 8pt;
                    color: #6c757d;
                 }

            }
            body {
                font-family: Arial, sans-serif; /* Fonte padrão */
                color: #333; /* Cor de texto principal */
                font-size: 10pt; /* Tamanho de fonte base */
                line-height: 1.4;
            }

            /* 1. Cabeçalho Oficial */
            .cabecalho-oficial {
                text-align: center;
                margin-bottom: 25px; /* Aumenta espaço após cabeçalho */
                border-bottom: 1px solid #ccc; /* Linha separadora sutil */
                padding-bottom: 10px;
            }
            .cabecalho-oficial p {
                margin: 2px 0;
                font-size: 10pt;
                color: #444; /* Cor mais escura que cinza padrão */
            }
            .cabecalho-oficial p:first-child { /* Título Principal */
                font-size: 12pt;
                font-weight: bold;
                color: #000;
                margin-bottom: 5px;
            }
             .cabecalho-oficial .periodo {
                font-weight: bold;
                margin-top: 8px;
            }

            /* Estilos gerais para Cards (seções) */
            .card {
                border: 1px solid #ccc; /* Borda mais sutil */
                border-radius: 0; /* Sem bordas arredondadas para visual oficial */
                margin-bottom: 15px;
                background-color: #fff;
                page-break-inside: avoid; /* Tenta evitar quebrar o card entre páginas */
            }
            .card-header {
                background-color: #e9ecef; /* Cinza claro neutro */
                color: #212529; /* Preto suave */
                padding: 6px 10px; /* Padding menor */
                border-bottom: 1px solid #ccc;
                font-weight: bold;
            }
            .card-header h2 {
                font-size: 11pt;
                margin: 0;
            }
            .card-body {
                padding: 10px;
            }
            .card-body p {
                margin-bottom: 6px;
            }

            /* 2. Seção de Identificação */
            .identificacao-section table {
                width: 100%;
                border-collapse: collapse;
                margin: 0; /* Remove margem padrão da tabela */
            }
            .identificacao-section td {
                padding: 3px 5px;
                border: none; /* Sem bordas internas na tabela de identificação */
                font-size: 9pt; /* Fonte ligeiramente menor */
            }
            .identificacao-section td.label {
                font-weight: bold;
                width: 150px; /* Largura fixa para os rótulos */
                color: #555;
            }

            /* 3. Resumo e Registros (Tabelas) */
            table { /* Estilo geral para tabelas de dados */
                width: 100%;
                border-collapse: collapse;
                margin: 10px 0;
                font-size: 9pt; /* Fonte menor para tabelas */
            }
            th, td {
                border: 1px solid #ccc; /* Bordas sutis */
                padding: 4px 6px; /* Padding ajustado */
                text-align: left;
                vertical-align: top;
            }
            th {
                background-color: #f2f2f2; /* Cinza muito claro para cabeçalhos */
                color: #333;
                font-weight: bold;
                text-align: center;
            }
            tr:nth-child(even) {
                 background-color: #f9f9f9; /* Fundo alternado muito sutil */
            }

            /* Ajustes específicos para tabela de registros */
            .registros-section th.col-data, .registros-section td.col-data { width: 70px; text-align: center; }
            .registros-section th.col-dia, .registros-section td.col-dia { width: 35px; text-align: center; }
            .registros-section th.col-hora, .registros-section td.col-hora { width: 55px; text-align: center; }
            .registros-section th.col-status, .registros-section td.col-status { width: 60px; text-align: center; }
            .registros-section th.col-detalhes, .registros-section td.col-detalhes {
                text-align: left;
                white-space: normal; /* Permite quebra de linha */
                word-wrap: break-word; /* Quebra palavras longas */
            }
             .registros-section td.col-detalhes strong {
                font-weight: bold;
                color: #555;
             }

            /* Estilos de Status (cores neutras) */
            .status-feriado { background-color: #eeeeee; color: #555; font-style: italic; }
            .status-fds { background-color: #f5f5f5; color: #777; font-style: italic; }
            .status-afastamento { background-color: #e0e0e0; color: #444; font-style: italic; }
            .status-pendente { background-color: #f5f5f5; color: #888; font-style: italic; }
            /* Saldo de horas */
            .saldo-horas { font-weight: bold; }
            .saldo-negativo { color: #c00; } /* Vermelho escuro sutil */


            /* 4. Seção de Autoavaliação */
            .autoavaliacao-section .card-header h2 { font-size: 10pt; }
            .autoavaliacao-section .card-body p {
                margin-left: 0; /* Remove indentação anterior */
                text-align: justify;
            }
            .declaracao-box {
                border: 1px solid #ccc;
                padding: 10px;
                margin-top: 15px;
                background-color: #f8f8f8;
            }
             .declaracao-box .card-header h2 { font-size: 10pt; }
             .declaracao-box .card-body p { margin-bottom: 10px; }
            .checkbox-simulado {
                 font-family: ZapfDingbats, sans-serif;
                 font-size: 11pt;
                 margin-right: 5px;
            }
            .assinatura {
                margin-top: 30px;
                text-align: center;
                color: #555;
                font-size: 9pt;
            }
            .assinatura-nome {
                text-align: center;
                font-weight: bold;
                margin-top: 5px;
                margin-bottom: 5px;
                 font-size: 10pt;
            }
            .data-assinatura {
                text-align: center;
                color: #555;
                 font-size: 9pt;
                 margin-bottom: 0;
            }

            /* Utilitários */
            .text-center { text-align: center; }
            .text-muted { color: #6c757d; }
            .fw-bold { font-weight: bold; }
            .fst-italic { font-style: italic; }
            .small { font-size: 9pt; } /* Ajustado para PDF */

        </style>
        """
        html_content = html_content.replace('</head>', f'{pdf_styles}</head>')
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "wb") as output_file:
            pisa_status = pisa.CreatePDF(html_content, dest=output_file)

        if pisa_status.err:
             logger.error(f"Erro do pisa ao gerar PDF: {pisa_status.err}")
             logger.error(f"Log do pisa: {pisa_status.log}")
             return False
        else:
             logger.info(f"PDF gerado com sucesso em: {output_path}")
             return True

    except Exception as e:
        # Usar o logger configurado em vez de current_app.logger se não estiver em contexto de app
        logger.error(f"Erro na função create_pdf: {e}", exc_info=True)
        return False

# Função create_excel (mantida como estava)
def create_excel(data, headers, output_path):
    """Cria um arquivo Excel a partir de uma lista de dicionários."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados" # Nome da planilha

        # Estilos (mantidos)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid") # Cinza claro
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Escreve o cabeçalho
        for col_idx, header_text in enumerate(headers.values(), 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escreve os dados
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers.keys(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, "")) # Usa .get para evitar erro se chave não existir
                cell.border = thin_border # Aplica borda a todas as células de dados

        # Ajusta largura das colunas (mantido)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15 # Largura padrão

        # Salva o arquivo
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel gerado com sucesso em: {output_path}")
        return True
    except Exception as e:
        logger.error(f"Erro ao criar Excel: {e}", exc_info=True)
        return False


# Função generate_pdf (ATUALIZADA para usar a função auxiliar importada)
def generate_pdf(user_id, mes, ano, context_completo=None):
    """
    Gera um arquivo PDF com o relatório mensal de ponto.
    Aceita um contexto completo opcional para incluir dados de autoavaliação.
    """
    try:
        context = {}
        # --- Usa a função auxiliar importada ---
        dados_base = _get_relatorio_mensal_data(user_id, mes, ano)
        # ---------------------------------------
        usuario = dados_base['usuario'] # Pega o usuário dos dados base

        if context_completo:
            # Se um contexto completo foi fornecido (para exportação do relatório salvo)
            # Apenas garante que os dados base estejam presentes (já estão em dados_base)
            context = {**dados_base, **context_completo} # Combina, garantindo dados base
            # Atualiza o título se necessário (pode já vir no context_completo)
            context['titulo'] = context.get('titulo', f'Relatório de Ponto e Autoavaliação - {dados_base["nome_mes"]}/{ano}')
        else:
            # Monta o contexto padrão (para PDF Padrão)
            context = {
                **dados_base, # Desempacota todos os dados base
                'data_geracao': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'titulo': f'Relatório de Ponto - {dados_base["nome_mes"]}/{ano}'
                # Não inclui chaves de autoavaliação aqui
            }

        # Define o caminho do arquivo de saída
        # Usa current_app se disponível, senão constrói caminho relativo
        try:
            static_folder = current_app.static_folder
        except RuntimeError: # Fora do contexto da aplicação
            # Tenta um caminho relativo (pode precisar de ajuste dependendo de onde é chamado)
            static_folder = os.path.join(os.path.dirname(__file__), '..', 'static')
            logger.warning("Executando generate_pdf fora do contexto da aplicação. Usando caminho relativo para static.")

        exports_dir = os.path.join(static_folder, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        # Adiciona sufixo '_completo' se for o relatório com autoavaliação
        filename_suffix = '_completo' if context_completo else ''
        filename = f"relatorio{filename_suffix}_{usuario.matricula}_{mes}_{ano}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        output_path = os.path.join(exports_dir, filename)

        # Cria o PDF usando o contexto montado
        success = create_pdf('exports/relatorio_ponto_pdf.html', output_path, **context)

        if success:
            # Retorna caminho relativo a partir da pasta static
            return f"exports/{filename}"
        else:
            raise Exception("Falha ao gerar PDF via create_pdf")
    except Exception as e:
        logger.error(f"Erro ao gerar PDF para user {user_id}, {mes}/{ano}: {e}", exc_info=True)
        return None

# Função generate_excel (ATUALIZADA para usar a função auxiliar importada)
def generate_excel(user_id, mes, ano):
    """Gera um arquivo Excel com o relatório mensal de ponto."""
    try:
        # Não precisa buscar usuário aqui, pois _get_relatorio_mensal_data já faz isso
        # --- Usa a função auxiliar importada ---
        dados_relatorio = _get_relatorio_mensal_data(user_id, mes, ano, order_desc=False)
        # ---------------------------------------
        usuario = dados_relatorio['usuario'] # Pega o usuário do resultado

        # Define os cabeçalhos do Excel
        headers = {
            'data': 'Data',
            'dia_semana': 'Dia da Semana',
            'entrada': 'Entrada',
            'saida_almoco': 'Saída Almoço',
            'retorno_almoco': 'Retorno Almoço',
            'saida': 'Saída',
            'horas_trabalhadas': 'Horas Trabalhadas',
            'status': 'Status', # Adicionado Status
            'observacoes': 'Observações',
            'resultados_produtos': 'Resultados/Produtos',
            'atividades': 'Atividades'
        }
        dias_semana_map = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
        data_excel = []

        # Itera sobre os dias do mês para incluir dias sem registro também
        for dia_num in range(1, dados_relatorio['ultimo_dia'].day + 1):
            data_atual = date(ano, mes, dia_num)
            dia_semana_idx = data_atual.weekday()
            registro = dados_relatorio['registros_por_data'].get(data_atual)
            is_feriado = data_atual in dados_relatorio['feriados_datas']
            is_fim_semana = dia_semana_idx >= 5

            row = {
                'data': data_atual.strftime('%d/%m/%Y'),
                'dia_semana': dias_semana_map[dia_semana_idx],
                'entrada': '', 'saida_almoco': '', 'retorno_almoco': '', 'saida': '',
                'horas_trabalhadas': '', 'status': '', 'observacoes': '',
                'resultados_produtos': '', 'atividades': ''
            }

            if is_feriado:
                row['status'] = f"Feriado ({dados_relatorio['feriados_dict'][data_atual]})"
            elif is_fim_semana and not registro:
                 row['status'] = "Fim de Semana" # Ou pode omitir
            elif registro:
                if registro.afastamento:
                    row['status'] = f"Afastamento ({registro.tipo_afastamento or 'N/A'})"
                    row['observacoes'] = registro.observacoes or ''
                    row['resultados_produtos'] = registro.resultados_produtos or ''
                else:
                    row['entrada'] = registro.entrada.strftime('%H:%M') if registro.entrada else ''
                    row['saida_almoco'] = registro.saida_almoco.strftime('%H:%M') if registro.saida_almoco else ''
                    row['retorno_almoco'] = registro.retorno_almoco.strftime('%H:%M') if registro.retorno_almoco else ''
                    row['saida'] = registro.saida.strftime('%H:%M') if registro.saida else ''
                    row['horas_trabalhadas'] = registro.horas_trabalhadas if registro.horas_trabalhadas is not None else ''
                    row['observacoes'] = registro.observacoes or ''
                    row['resultados_produtos'] = registro.resultados_produtos or ''
                    # Busca atividades
                    lista_atividades = dados_relatorio['atividades_por_ponto'].get(registro.id, [])
                    row['atividades'] = "; ".join(lista_atividades) if lista_atividades else ''
                    # Define status baseado nas horas
                    if registro.horas_trabalhadas is not None:
                         row['status'] = 'OK' if registro.horas_trabalhadas >= 8 else 'Parcial'
                    else:
                         row['status'] = 'Pendente'
            elif not is_fim_semana: # Dia útil sem registro
                row['status'] = 'Pendente (Sem Registro)'

            data_excel.append(row)

        # Define caminho e nome do arquivo
        # Usa current_app se disponível, senão constrói caminho relativo
        try:
            static_folder = current_app.static_folder
        except RuntimeError: # Fora do contexto da aplicação
            static_folder = os.path.join(os.path.dirname(__file__), '..', 'static')
            logger.warning("Executando generate_excel fora do contexto da aplicação. Usando caminho relativo para static.")

        exports_dir = os.path.join(static_folder, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        filename = f"relatorio_{usuario.matricula}_{mes}_{ano}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        output_path = os.path.join(exports_dir, filename)

        # Cria o arquivo Excel
        success = create_excel(data_excel, headers, output_path)

        if success:
            # Retorna caminho relativo a partir da pasta static
            return f"exports/{filename}"
        else:
            raise Exception("Falha ao gerar Excel")
    except Exception as e:
        logger.error(f"Erro ao gerar Excel para user {user_id}, {mes}/{ano}: {e}", exc_info=True)
        return None

# Aliases antigos (mantidos por compatibilidade, mas podem ser removidos)
export_registros_pdf = generate_pdf
export_registros_excel = generate_excel
